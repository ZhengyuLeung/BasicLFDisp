import random

import numpy as np
import torch
from pathlib import Path
import logging
from option import args
from einops import rearrange
import openpyxl
import torch.nn.functional as F
from skimage.metrics import peak_signal_noise_ratio as cal_PSNR
from skimage.metrics import structural_similarity as cal_SSIM


class ExcelFile(object):
    def __init__(self):
        self.xlsx_file = openpyxl.Workbook()
        self.worksheet = self.xlsx_file.active
        self.worksheet.title = 'sheet1'
        self.header_list = ['Datasets', 'Scenes', 'MSE_100', 'BP_07', 'BP_03', 'BP_01']

        self.sum = 1
        self.worksheet.cell(self.sum, 1, 'Datasets')
        self.worksheet.cell(self.sum, 2, 'Scenes')
        self.worksheet.column_dimensions['A'].width = 16
        self.worksheet.column_dimensions['B'].width = 22
        self.add_count(1)

    def write_sheet(self, test_name, LF_name, metric_name, metric_score):
        self.worksheet.cell(self.sum, 1, test_name)
        self.worksheet.cell(self.sum, 2, LF_name)

        # self.worksheet.col(self.header_list.index(metric_name)).width = 256 * 10
        self.worksheet.cell(1, self.header_list.index(metric_name)+1, metric_name)
        self.worksheet.cell(self.sum, self.header_list.index(metric_name)+1, '%.6f' % metric_score)

    def add_count(self, num):
        self.sum = self.sum + num


def get_logger(log_dir, args):
    '''LOG '''
    logger = logging.getLogger(args.model_name)
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler = logging.FileHandler('%s/%s.txt' % (log_dir, args.model_name))
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def create_dir(args):
    log_dir = Path(args.path_log)
    log_dir.mkdir(exist_ok=True)
    if args.task == 'SSR':
        task_path = args.task + '_' + str(args.scale_factor) + 'x'
    elif args.task == 'BlindSSR':
        task_path = args.task + '_' + str(args.scale_factor) + 'x' + '_sigma_%.1f_noise_%.1f' % (args.sig, args.noise)
    elif args.task == 'ASR':
        task_path = 'ASR_' + str(args.angRes_in) + 'x' + str(args.angRes_in) + '_' + str(args.angRes_out) + 'x' + str(args.angRes_out)
    elif args.task == 'Disp':
        task_path = 'Disp' + '_' + str(args.angRes_in) + 'x' + str(args.angRes_in)

    log_dir = log_dir.joinpath(task_path)
    log_dir.mkdir(exist_ok=True)
    log_dir = log_dir.joinpath(args.model_name)
    log_dir.mkdir(exist_ok=True)

    checkpoints_dir = log_dir.joinpath('checkpoints/')
    checkpoints_dir.mkdir(exist_ok=True)

    results_dir = log_dir.joinpath('results/')
    results_dir.mkdir(exist_ok=True)

    return log_dir, checkpoints_dir, results_dir


class Logger():
    def __init__(self, log_dir, args):
        self.logger = get_logger(log_dir, args)

    def log_string(self, str):
        if args.local_rank <= 0:
            self.logger.info(str)
            print(str)


def cal_metrics(label, out, bd=0):
    [_, _, H, W] = label.size()
    label = (label.data.squeeze().cpu())[bd:H-bd, bd:W-bd]
    out = (out.data.squeeze().cpu())[bd:H-bd, bd:W-bd]

    # mask_valid = ~(torch.isnan(label) + torch.isneginf(label) + torch.isposinf(label) + \
    #                torch.isnan(out) + torch.isneginf(out) + torch.isposinf(out))

    diff = torch.abs(out - label)
    mse_x100 = 100 * torch.mean(torch.square(diff))  #[mask_valid])
    bad_pixel_07 = 100 * torch.mean(torch.where(diff >= 0.07, 1, 0).float())
    bad_pixel_03 = 100 * torch.mean(torch.where(diff >= 0.03, 1, 0).float())
    bad_pixel_01 = 100 * torch.mean(torch.where(diff >= 0.01, 1, 0).float())

    return mse_x100, bad_pixel_07, bad_pixel_03, bad_pixel_01


class LF_divide_integrate(object):
    def __init__(self, scale, patch_size, stride):
        self.scale = scale
        self.patch_size = patch_size
        self.stride = stride
        self.bdr = (patch_size - stride) // 2
        self.pad = torch.nn.ReflectionPad2d(padding=(self.bdr, self.bdr + stride - 1, self.bdr, self.bdr + stride - 1))

    def LFdivide(self, LF):
        assert LF.size(0) == 1, 'The batch_size of LF for test requires to be one!'
        LF = LF.squeeze(0)
        [c, u, v, h, w] = LF.size()

        LF = rearrange(LF, 'c u v h w -> (c u v) 1 h w')
        self.numU = (h + self.bdr * 2 - 1) // self.stride
        self.numV = (w + self.bdr * 2 - 1) // self.stride

        # LF_pad = self.pad(LF)
        LF_pad = ImageExtend(LF, [self.bdr, self.bdr + self.stride - 1, self.bdr, self.bdr + self.stride - 1])
        LF_divided = F.unfold(LF_pad, kernel_size=self.patch_size, stride=self.stride)
        LF_divided = rearrange(LF_divided, '(c u v) (h w) (numU numV) -> (numU numV) c u v h w', u=u, v=v,
                               h=self.patch_size, w=self.patch_size, numU=self.numU, numV=self.numV)
        return LF_divided

    def LFintegrate(self, LF_divided):
        LF_divided = LF_divided[:, :, self.bdr*self.scale:(self.bdr+self.stride)*self.scale,
                                self.bdr*self.scale:(self.bdr+self.stride)*self.scale]
        LF = rearrange(LF_divided, '(numU numV) c h w -> c (numU h) (numV w)', numU=self.numU, numV=self.numV)
        return LF


def ImageExtend(Im, bdr):
    [_, _, h, w] = Im.size()
    Im_lr = torch.flip(Im, dims=[-1])
    Im_ud = torch.flip(Im, dims=[-2])
    Im_diag = torch.flip(Im, dims=[-1, -2])

    Im_up = torch.cat((Im_diag, Im_ud, Im_diag), dim=-1)
    Im_mid = torch.cat((Im_lr, Im, Im_lr), dim=-1)
    Im_down = torch.cat((Im_diag, Im_ud, Im_diag), dim=-1)
    Im_Ext = torch.cat((Im_up, Im_mid, Im_down), dim=-2)
    Im_out = Im_Ext[:, :, h - bdr[0]: 2 * h + bdr[1], w - bdr[2]: 2 * w + bdr[3]]
    return Im_out


def LF_rgb2ycbcr(x):
    y = torch.zeros_like(x)
    y[:,0,:,:,:,:] =  65.481 * x[:,0,:,:,:,:] + 128.553 * x[:,1,:,:,:,:] +  24.966 * x[:,2,:,:,:,:] +  16.0
    y[:,1,:,:,:,:] = -37.797 * x[:,0,:,:,:,:] -  74.203 * x[:,1,:,:,:,:] + 112.000 * x[:,2,:,:,:,:] + 128.0
    y[:,2,:,:,:,:] = 112.000 * x[:,0,:,:,:,:] -  93.786 * x[:,1,:,:,:,:] -  18.214 * x[:,2,:,:,:,:] + 128.0

    y = y / 255.0
    return y


def LF_ycbcr2rgb(x):
    mat = np.array(
        [[65.481, 128.553, 24.966],
         [-37.797, -74.203, 112.0],
         [112.0, -93.786, -18.214]])
    mat_inv = np.linalg.inv(mat)
    offset = np.matmul(mat_inv, np.array([16, 128, 128]))
    mat_inv = mat_inv * 255

    y = torch.zeros_like(x)
    y[:,0,:,:,:,:] = mat_inv[0,0] * x[:,0,:,:,:,:] + mat_inv[0,1] * x[:,1,:,:,:,:] + mat_inv[0,2] * x[:,2,:,:,:,:] - offset[0]
    y[:,1,:,:,:,:] = mat_inv[1,0] * x[:,0,:,:,:,:] + mat_inv[1,1] * x[:,1,:,:,:,:] + mat_inv[1,2] * x[:,2,:,:,:,:] - offset[1]
    y[:,2,:,:,:,:] = mat_inv[2,0] * x[:,0,:,:,:,:] + mat_inv[2,1] * x[:,1,:,:,:,:] + mat_inv[2,2] * x[:,2,:,:,:,:] - offset[2]
    return y


def LF_interpolate(LF, size=None, scale_factor=None, mode='nearest'):
    [b, c, u, v, h, w] = LF.size()
    LF = rearrange(LF, 'b c u v h w -> (b u v) c h w')
    if size is None:
        size = [h*scale_factor, w*scale_factor]
    if mode == 'nearest':
        LF_upscale = F.interpolate(LF, size=size, mode=mode)
    else:
        LF_upscale = F.interpolate(LF, size=size, mode=mode, align_corners=False)
    LF_upscale = rearrange(LF_upscale, '(b u v) c h w -> b c u v h w', u=u, v=v)
    return LF_upscale


def LF_interpolate_ang(LF, size, mode, align_corners=False):
    [b, c, u, v, h, w] = LF.size()
    LF = rearrange(LF, 'b c u v h w -> (b h w) c u v')
    if mode == 'nearest':
        LF_upscale = F.interpolate(LF, size=size, mode=mode)
    else:
        LF_upscale = F.interpolate(LF, size=size, mode=mode, align_corners=align_corners)
    LF_upscale = rearrange(LF_upscale, '(b h w) c u v -> b c u v h w', h=h, w=w)
    return LF_upscale


def LF_random_crop_ang(LF, angRes_U_req, angRes_V_req):
    def gen_index(angRes, angRes_req):
        if angRes_req == 1:
            stride = 1000
            start = random.randint(0, angRes_req - 1)
        else:
            stride = random.randint(1, int((angRes - 1) / (angRes_req - 1)))
            start = random.randint(0, angRes - stride * (angRes_req - 1) - 1)
        return [start, stride]

    [_, _, angRes_U, angRes_V, _, _] = LF.size()
    [start_U, stride_U] = gen_index(angRes_U, angRes_U_req)
    [start_V, stride_V] = gen_index(angRes_V, angRes_V_req)
    LF = LF[:, :, start_U:angRes_U:stride_U, start_V:angRes_V:stride_V, :, :]
    return LF


def read_pfm(fpath, expected_identifier="Pf"):
    # PFM format definition: http://netpbm.sourceforge.net/doc/pfm.html

    def _get_next_line(f):
        next_line = f.readline().decode('utf-8').rstrip()
        # ignore comments
        while next_line.startswith('#'):
            next_line = f.readline().rstrip()
        return next_line

    with open(fpath, 'rb') as f:
        #  header
        identifier = _get_next_line(f)
        if identifier != expected_identifier:
            raise Exception('Unknown identifier. Expected: "%s", got: "%s".' % (expected_identifier, identifier))

        try:
            line_dimensions = _get_next_line(f)
            dimensions = line_dimensions.split(' ')
            width = int(dimensions[0].strip())
            height = int(dimensions[1].strip())
        except:
            raise Exception('Could not parse dimensions: "%s". '
                            'Expected "width height", e.g. "512 512".' % line_dimensions)

        try:
            line_scale = _get_next_line(f)
            scale = float(line_scale)
            assert scale != 0
            if scale < 0:
                endianness = "<"
            else:
                endianness = ">"
        except:
            raise Exception('Could not parse max value / endianess information: "%s". '
                            'Should be a non-zero number.' % line_scale)

        try:
            data = np.fromfile(f, "%sf" % endianness)
            data = np.reshape(data, (height, width))
            data = np.flipud(data)
            with np.errstate(invalid="ignore"):
                data *= abs(scale)
        except:
            raise Exception('Invalid binary values. Could not create %dx%d array from input.' % (height, width))

        return data


def LF_Bicubic(LF, scale=1/4):
    [b, c, u, v, h, w] = LF.size()
    LF = rearrange(LF, 'b c u v h w -> (b u v) c h w')
    LF_bic = Bicubic()(LF, scale)
    LF_bic = rearrange(LF_bic, '(b u v) c h w -> b c u v h w', u=u, v=v)
    return LF_bic


# implementation of matlab bicubic interpolation in pytorch
class Bicubic(object):
    def __init__(self):
        super(Bicubic, self).__init__()

    def cubic(self, x):
        absx = torch.abs(x)
        absx2 = torch.abs(x) * torch.abs(x)
        absx3 = torch.abs(x) * torch.abs(x) * torch.abs(x)

        condition1 = (absx <= 1).to(torch.float32)
        condition2 = ((1 < absx) & (absx <= 2)).to(torch.float32)

        f = (1.5 * absx3 - 2.5 * absx2 + 1) * condition1 + (-0.5 * absx3 + 2.5 * absx2 - 4 * absx + 2) * condition2
        return f

    def contribute(self, in_size, out_size, scale):
        kernel_width = 4
        if scale < 1:
            kernel_width = 4 / scale
        x0 = torch.arange(start=1, end=out_size[0] + 1).to(torch.float32)
        x1 = torch.arange(start=1, end=out_size[1] + 1).to(torch.float32)

        u0 = x0 / scale + 0.5 * (1 - 1 / scale)
        u1 = x1 / scale + 0.5 * (1 - 1 / scale)

        left0 = torch.floor(u0 - kernel_width / 2)
        left1 = torch.floor(u1 - kernel_width / 2)

        P = np.ceil(kernel_width) + 2

        indice0 = left0.unsqueeze(1) + torch.arange(start=0, end=P).to(torch.float32).unsqueeze(0)
        indice1 = left1.unsqueeze(1) + torch.arange(start=0, end=P).to(torch.float32).unsqueeze(0)

        mid0 = u0.unsqueeze(1) - indice0.unsqueeze(0)
        mid1 = u1.unsqueeze(1) - indice1.unsqueeze(0)

        if scale < 1:
            weight0 = scale * self.cubic(mid0 * scale)
            weight1 = scale * self.cubic(mid1 * scale)
        else:
            weight0 = self.cubic(mid0)
            weight1 = self.cubic(mid1)

        weight0 = weight0 / (torch.sum(weight0, 2).unsqueeze(2))
        weight1 = weight1 / (torch.sum(weight1, 2).unsqueeze(2))

        indice0 = torch.min(torch.max(torch.FloatTensor([1]), indice0), torch.FloatTensor([in_size[0]])).unsqueeze(0)
        indice1 = torch.min(torch.max(torch.FloatTensor([1]), indice1), torch.FloatTensor([in_size[1]])).unsqueeze(0)

        kill0 = torch.eq(weight0, 0)[0][0]
        kill1 = torch.eq(weight1, 0)[0][0]

        weight0 = weight0[:, :, kill0 == 0]
        weight1 = weight1[:, :, kill1 == 0]

        indice0 = indice0[:, :, kill0 == 0]
        indice1 = indice1[:, :, kill1 == 0]

        return weight0, weight1, indice0, indice1

    def __call__(self, input, scale=1/4):
        b, c, h, w = input.shape

        weight0, weight1, indice0, indice1 = self.contribute([h, w], [int(h * scale), int(w * scale)], scale)
        weight0 = weight0[0].to(input.device)
        weight1 = weight1[0].to(input.device)

        indice0 = indice0[0].long()
        indice1 = indice1[0].long()

        out = input[:, :, (indice0 - 1), :] * (weight0.unsqueeze(0).unsqueeze(1).unsqueeze(4))
        out = torch.sum(out, dim=3)
        A = out.permute(0, 1, 3, 2)

        out = A[:, :, (indice1 - 1), :] * (weight1.unsqueeze(0).unsqueeze(1).unsqueeze(4))
        out = out.sum(3).permute(0, 1, 3, 2)

        return out